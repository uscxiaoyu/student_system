import re
import os, sys
import django
from xlrd import open_workbook
from openpyxl import load_workbook


PROJ_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(PROJ_ROOT)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "student_sys.settings")
django.setup()

from student.models import Student, Project, StudentJoinProject

BASE_DIR = "data/"
def load_student():
    return
    
def load_project():
    return


def load_excel(f_path, col_names=["学号", "姓名", "性别", "学院", "年级", "专业", "活动类型", "活动时间", "主办单位", "活动名称", "认证状态"]):
    d_dict = {}
    if '.xlsx' in f_path:
        wb = load_workbook(filename=f_path)
        sheet = wb[wb.sheetnames[0]]
        for col in sheet.columns:
            col_value = [re.sub("\s", "", c.value) if isinstance(c.value, str) else c.value for c in col]
            for v in col_names:
                if v in col_value:
                    d_dict[v] = col_value[col_value.index(v)+1: ]
                
    elif '.xls' in f_path:
        wb = open_workbook(filename=f_path)
        sheet = wb.sheet_by_index(0)
        for i in range(sheet.ncols):
            col_value = [re.sub("\s", "", c.value) if isinstance(c.value, str) else c.value for c in sheet.col(i)]
            for v in col_names:
                if v in col_value:
                    d_dict[v] = col_value[col_value.index(v)+1: ]
    else:
        print("请输入xls或者xlsx文件!")
                
    return d_dict


def load_studentjoinproject(dir_name):
    f_paths = os.listdir(dir_name)  # 读取所有文件名
    pattern = re.compile("\d+\s+(.+).xls")  # 匹配文件名
    try:
        f = open(dir_name + "/log.txt", "w")
        print(f"开始导入{dir_name}文件夹中的excel文件:")
        for f_path in f_paths:
            prj_names = pattern.findall(f_path)
            if prj_names:  # 如果是合法的文件
                try:
                    d_dict= load_excel(dir_name + '/' + f_path)
                    prj_name = prj_names[0]
                    for i, s_id in enumerate(d_dict["学号"]):
                        if isinstance(s_id, (int, float)):
                            StudentJoinProject.objects.create(s_id=s_id, student_name=d_dict["姓名"][i], project_name=prj_name)
                        elif isinstance(s_id, str) and ('/' in s_id or 'EBI' in s_id):
                            StudentJoinProject.objects.create(s_id=s_id, student_name=d_dict["姓名"][i], project_name=prj_name)
                        else:
                            f.write(f_path + '    ' + str(s_id) + '  '+ str(d_dict["姓名"][i]) + '\n')
                    print("  ", f_path, "导入成功")
                except Exception as e:
                    print("  ", f_path, "导入失败")
                    f.write(f_path + "出错: " + str(e) + '\n')
        f.write("\n\n")
    except Exception as e:
        f.write("出错: " + str(e) + '\n')
    finally:
        f.close()
    
    print(dir_name, "导入完成!")
    return

if __name__ == "__main__":
    # student = Student.objects.all()[0]
    # print(student.student_id, student.name, student.department_name)
    s_dir= BASE_DIR + "活动名称与学生匹配表/"
    dirs = [s_dir+d for d in os.listdir(s_dir)]
    for dir_name in dirs:
        load_studentjoinproject(dir_name)
        