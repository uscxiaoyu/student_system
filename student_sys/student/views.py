import enum
from openpyxl import load_workbook
from django.contrib import auth
from django.http import JsonResponse
from django.http.response import Http404, HttpResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect
from django.contrib.auth.models import User, Permission, Group
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse
from django.views import View
from .models import Student, Project, StudentJoinProject, StudentOrganization, StudentScholar, UserProfile
from workflow.models import ScholarProcess, OrganizationProcess
from .write_docx import ReportDocx
from .access_databases import getOrgByDeptName, getScholarByDeptName
import json
import datetime
import re


class LoginView(View):
    login_name = "login.html"
    teacher_main_page = "teacher_main.html"
    student_main_page = "student_main.html"

    def get(self, request):
        is_login = request.session.get("is_login", False)
        if is_login:
            u = request.session["username"]
            role = request.session["role"]
            print(u, role)
            if role == "教师":
                request.session["role"] = "教师"
                user = User.objects.get(username=u)
                try:
                    u_department = user.userprofile.get_department_display()  # 教师所在部门
                    org_res = getOrgByDeptName(dept_name=u_department)
                    scholar_res = getScholarByDeptName(dept_name=u_department)
                    print(u_department, org_res, scholar_res)
                    return render(
                        request,
                        self.teacher_main_page,
                        {"user": u, "org_res": org_res, "scholar_res": scholar_res},
                    )
                except Exception as e:
                    print(e)
                    return render(
                        request,
                        self.teacher_main_page,
                        {"user": u, "org_res": [], "scholar_res": []},
                    )
            else:
                reportDoc = ReportDocx(u)
                infos = reportDoc.hint_list  # 错误提示
                report = reportDoc.student_report  # 学生信息
                return render(
                    request,
                    self.student_main_page,
                    {
                        "user": u,
                        "infos": infos,
                        "meta_info": report["基础信息"],
                        "activities": report["参加活动经历"],
                        "organization": report["学生组织经历"],
                        "scholar": report["获奖经历"],
                    },
                )
        else:
            return render(request, self.login_name)

    def post(self, request):
        u = request.POST.get("_id")  # 学号或者教工号
        p = request.POST.get("pwd")
        if User.objects.filter(username=u):
            user = User.objects.get(username=u)
            u_id = user.id
            u_group = Group.objects.get(user=u_id).name
            user = authenticate(username=u, password=p)
            if user:
                if user.is_active:
                    login(request, user)  # 登陆用户
                    # 在session中记录用户和角色
                    request.session["is_login"] = True
                    request.session["username"] = u
                    if u_group == "教师":
                        request.session["role"] = "教师"
                        try:
                            u_department = user.userprofile.get_department_display()  # 教师所在部门
                            org_res = getOrgByDeptName(dept_name=u_department)
                            scholar_res = getScholarByDeptName(dept_name=u_department)
                            return render(
                                request,
                                self.teacher_main_page,
                                {"request": request, "user": u, "org_res": org_res, "scholar_res": scholar_res},
                            )
                        except Exception as e:
                            print(e)
                            return render(
                                request,
                                self.teacher_main_page,
                                {"request": request, "user": u, "org_res": [], "scholar_res": []},
                            )
                    else:
                        request.session["role"] = "学生"
                        reportDoc = ReportDocx(u)
                        infos = reportDoc.hint_list  # 错误提示
                        report = reportDoc.student_report  # 学生信息
                        return render(
                            request,
                            self.student_main_page,
                            {
                                "request": request,
                                "user": u,
                                "infos": infos,
                                "meta_info": report["基础信息"],
                                "activities": report["参加活动经历"],
                                "organization": report["学生组织经历"],
                                "scholar": report["获奖经历"],
                            },
                        )
                else:
                    self.tips = "账号未激活，请联系管理员"
            else:
                self.tips = "帐号密码错误，请重新输入"
        else:
            self.tips = "用户不存在，请注册"

        return render(request, self.login_name)


def logoutView(request):
    logout(request)
    return redirect("student:login")


def signUpView(request):
    if request.method == "GET":
        return render(request, "signUp.html")

    if request.method == "POST":
        body = request.body.decode("utf-8")
        res = json.loads(body)
        username = res["username"]
        password = res["password"]
        email = res["email"]
        print(res)
        if User.objects.filter(username=username):
            return HttpResponse(f"<span style='color: red'>用户{username}已有注册记录！若非本人注册，请联系辅导员！</span>")
        else:
            try:
                user = User()
                user.username = username
                user.set_password(password)  # 注意，此处会转换，数据库不保存明文密码
                user.email = email
                user.is_active = 1
                user.is_staff = 0
                user.date_joined = datetime.datetime.now()
                user.save()
                user.groups.add(2)
                return HttpResponse("<span>恭喜你！注册成功！<a href='/'>返回登陆页</a></span>")
            except Exception as e:
                return HttpResponse(str(e))


def downloadDocxView(request):
    if request.method == "POST":  # 教师页面发出
        student_id = request.POST.get("_id")
    else:  # 学生页面发出
        student_id = request.session["username"]

    reportDoc = ReportDocx(student_id)
    reportDoc.generate_docx()  # 生成docx文件
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
    response["Content-Disposition"] = f"attachment; filename={student_id}-report.docx"
    reportDoc.document.save(response)

    return response


@login_required(login_url="")
def checkDocxView(request):
    body = request.body.decode("utf-8")
    res = json.loads(body)
    student_id = res["s_id"]
    reportDoc = ReportDocx(student_id)
    # infos = reportDoc.hint_list
    report = reportDoc.student_report
    response = ""
    if not report:
        response += f"<p>学号{student_id}不存在！</p>"
        return HttpResponse(response)
    else:
        response = ""
        base_info = """
            <h3>个人基本信息</h3>
            <table class='layui-table' lay-skin='line'>
                <tr>
                    <th>所在学院</th><th>年级</th><th>专业</th><th>学号</th><th>姓名</th><th>性别</th>
                </tr>
                <tr>
                    <td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>
                </tr>
            </table>
            """ % tuple(
            report["基础信息"]
        )
        response += base_info
        head = "<tr> <th>%s</th> <th>%s</th> <th>%s</th> <th>%s</th> <th>%s</th></tr>" % (
            "序号",
            "活动日期",
            "活动名称",
            "主办单位",
            "认证状态",
        )

        for cate in report["参加活动经历"]:
            prj_html = f"<h3>{cate}</h3>"
            table = "<table class='layui-table' lay-skin='line'>" + head
            for prj in report["参加活动经历"][cate]:
                row = "<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % tuple(prj)
                table += row

            table += "</table>"
            prj_html += table
            response += prj_html

        if report["学生组织经历"]:
            prj_html = f"<h3> 学生组织经历 </h3>"
            table = """<table class='layui-table' lay-skin='line'>
                <tr>
                    <th>序号</th><th>开始时间</th><th>结束时间</th><th>所在部门</th><th>职位</th>
                    <th>指导单位</th><th>认证状态</th>
                </tr>
            """
            for prj in report["学生组织经历"]:
                row = "<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % tuple(
                    prj[:-1]
                )
                table += row

            table += "</table>"
            prj_html += table
            response += prj_html

        if report["获奖经历"]:
            prj_html = f"<h3>获奖经历</h3>"
            table = """<table class='layui-table' lay-skin='line'>
                <tr>
                    <th>序号</th><th>获奖时间</th><th>奖励名称</th><th>奖励级别</th><th>认证状态</th>
                </tr>
            """
            for prj in report["获奖经历"]:
                row = "<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>" % tuple(prj[:-1])
                table += row

            table += "</table>"
            prj_html += table
            response += prj_html

    # if infos:
    #     response += "<p>以下项目查询出错，很可能是没在项目表中找到对应的项目</p>"
    #     for i, info in enumerate(infos):
    #         a = "<p>" + f"{i+1}. " + info[0] + ": " + info[1] + "</p>"
    #         response += a

    return HttpResponse(response)


def insertStudentScholar(request):
    body = request.body.decode("utf-8")
    res = json.loads(body)
    if "s_id" not in res:
        res["s_id"] = request.user
    try:
        r = StudentScholar.objects.create(**res)
        return JsonResponse({"id": r.id})
    except Exception as e:
        print(e)
        return HttpResponse("添加失败" + str(e))


@login_required(login_url="")
def deleteStudentScholar(request, r_id):
    try:
        StudentScholar.objects.get(id=r_id).delete()
        return HttpResponse(f"删除学生组织经历{r_id}成功")
    except Exception as e:
        return HttpResponse("删除失败")


def insertStudentOrganization(request):
    body = request.body.decode("utf-8")
    res = json.loads(body)
    if "s_id" not in res:
        res["s_id"] = request.user
    try:
        StudentOrganization.objects.create(**res)
        print(res, "插入成功!")
        return HttpResponse("<p>添加成功</p>")
    except Exception as e:
        print(e)
        return HttpResponse("添加失败" + str(e))


@login_required(login_url="")
def deleteStudentOrganization(request, r_id):
    try:
        StudentOrganization.objects.get(id=r_id).delete()
        return HttpResponse(f"删除获奖经历{r_id}成功")
    except Exception as e:
        return HttpResponse("删除失败")


@login_required(login_url="")
def updateOrganizationState(request):
    try:
        body = request.body.decode("utf-8")
        user = request.session["username"]
        res = json.loads(body)
        org_id = res["_id"]
        state = res["state"]
        inst_org = StudentOrganization.objects.get(id=org_id)
        inst_org.certify_state = state
        OrganizationProcess.objects.create(
            **{
                "student_id": inst_org.s_id,
                "org_id": org_id,
                "admin_id": user,
                "update_time": datetime.datetime.now(),
                "state": state,
            }
        )
        return HttpResponse("更新成功")
    except Exception as e:
        print(e)
        return HttpResponse("更新失败")


@login_required(login_url="")
def updateScholarState(request):
    try:
        body = request.body.decode("utf-8")
        user = request.session["username"]
        res = json.loads(body)
        scholar_id = res["_id"]
        state = res["state"]
        inst_scholar = StudentScholar.objects.get(id=scholar_id)
        inst_scholar.certify_state = state
        ScholarProcess.objects.create(
            **{
                "student_id": inst_scholar.s_id,
                "scholar_id": scholar_id,
                "admin_id": user,
                "update_time": datetime.datetime.now(),
                "state": state,
            }
        )
        print(inst_scholar)
        return HttpResponse("更新成功")
    except Exception as e:
        print(e)
        return HttpResponse("更新失败")


@login_required(login_url="")
def updateStudentOrganization(request):
    try:
        body = request.body.decode("utf-8")
        res = json.loads(body)
        _id = res["_id"]
        org = StudentOrganization.objects.get(id=_id)
        org.start_time = res["start_time"]
        org.end_time = res["end_time"]
        org.org_name = res["org_name"]
        org.position = res["position"]
        org.department_name = res["department_name"]
        org.update_time = datetime.datetime.now()
        org.certify_state = "未认证"  # 修改之后默认重新提交
        org.save()
        return HttpResponse("更新成功")
    except Exception as e:
        return HttpResponse(str(e))


@login_required(login_url="")
def updateStudentScholar(request):
    try:
        body = request.body.decode("utf-8")
        res = json.loads(body)
        _id = res["_id"]
        scholar = StudentScholar.objects.get(id=_id)
        scholar.g_time = res["g_time"]
        scholar.scholar_name = res["scholar_name"]
        scholar.level = res["level"]
        scholar.certify_state = "未认证"  # 修改之后默认重新提交
        scholar.update_time = datetime.datetime.now()
        scholar.save()
        return HttpResponse("更新成功")
    except Exception as e:
        return HttpResponse(str(e))


@login_required(login_url="")
def uploadFile(request):
    """
    上传学生奖学金记录或组织活动记录
    """
    try:
        col_names = ["学号", "奖学金名称", "奖学金级别", "组织名称", "职位", "隶属部门", "开始时间", "结束时间", "获奖时间"]
        print(request.FILES["file"])
        wb = load_workbook(request.FILES["file"])
        sheet = wb[wb.sheetnames[0]]
        d_dict = {}
        for col in sheet.columns:
            col_value = [re.sub("\s", "", c.value) if isinstance(c.value, str) else c.value for c in col]
            for v in col_names:
                if v in col_value:
                    d_dict[v] = col_value[col_value.index(v) + 1 :]
        print(d_dict)
        if len(d_dict) == 4:
            for i, s_id in enumerate(d_dict["学号"]):
                s = StudentScholar()
                s.s_id = s_id
                s.g_time = d_dict["获奖时间"][i]
                s.scholar_name = d_dict["奖学金名称"][i]
                s.level = d_dict["奖学金级别"][i]
                s.certify_state = "已认证"
                s.save()
        elif len(d_dict) == 6:
            for i, s_id in enumerate(d_dict["学号"]):
                o = StudentOrganization()
                o.s_id = s_id
                o.org_name = d_dict["组织名称"][i]
                o.position = d_dict["职位"][i]
                o.start_time = d_dict["开始时间"][i]
                o.end_time = d_dict["结束时间"][i]
                o.department_name = d_dict["部门名称"][i]
                o.certify_state = "已认证"
                o.save()
        else:
            raise ValueError("表格列数不对！")

        return JsonResponse({"res": "上传记录成功!"})
    except Exception as e:
        print(e)
        return JsonResponse({"res": "上传记录失败!", "error": str(e)})